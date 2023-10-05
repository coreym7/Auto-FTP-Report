import os
import paramiko
import datetime
from ftplib import FTP, error_perm
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import time
import logging
#import requests

# Setup the logger at the beginning of your script
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# Create a file handler
handler = logging.FileHandler('logfile.log')

# Create a logging format
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)

# Add the handlers to the logger
logger.addHandler(handler)

logger.info("Script has started.")  # New logging line

# FTP Credentials
"""
details for FTP would be implemented here
"""

remote_file_path = "remote/file/path.csv"
local_file_path = "path/to/downloaded/report.csv"

current_date = datetime.datetime.now().strftime('%Y-%m-%d')
local_file_with_date = local_file_path.replace('.csv', f'_{current_date}.csv')
xlsx_file_path = local_file_with_date.replace('.csv', '.xlsx')
YOUR_URL = "https://ftpURL.com"

def download_report():
    try:
        # Create a transport object
        transport = paramiko.Transport((sftp_host, sftp_port))
        
        # Authenticate the transport object
        transport.connect(username=sftp_username, password=sftp_password)
        
        # Create an SFTP client from the transport object
        sftp = paramiko.SFTPClient.from_transport(transport)
                
        # Check if the remote file exists
        file_stat = sftp.stat(remote_file_path)
        if file_stat:
            # Retrieve the last modified time of the file
            last_modified_time = file_stat.st_mtime

            # Get the current time
            current_time = time.time()

            # Check if the file was modified in the last 24 hours
            if current_time - last_modified_time <= 24 * 60 * 60:
                # Download the file
                sftp.get(remote_file_path, local_file_with_date)
                logger.info(f"Report downloaded: {local_file_with_date}")
                return True
            else:
                logger.error("Report not found or not modified in the last 24 hours.")
                return False
        else:
            logger.error("Report not found.")
            return False
            # Here's the fallback placement:
            """
            if not download_file_from_url(YOUR_URL, local_file_with_date):
                logger.error("Both primary and fallback download methods failed.")
                return False
            return True
            """


    except Exception as e:
        logger.error(f"An error occurred while downloading the report: {e}")
    
    finally:
        # Close the SFTP client and the transport object
        sftp.close()
        transport.close()


def csv_to_excel(csv_path, excel_path):
    try:
        # Read the CSV data
        df = pd.read_csv(csv_path)

        # Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Write the DataFrame to the worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Adjust the columns widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # Save the workbook
        wb.save(excel_path)
        logger.info(f"Excel file saved: {excel_path}")

    except Exception as e:
        logger.error(f"An error occurred while converting csv to excel: {e}")

if __name__ == "__main__":
    try:
        download_success = download_report()
        # Only process csv if download was successful
        if download_success:
            csv_to_excel(local_file_with_date, xlsx_file_path)
        else:
            logger.error("Download failed. Skipping CSV to Excel conversion.")
        
        # Delete local download file to prevent errors in future run and to keep download folder clean    
        if os.path.exists (local_file_with_date):
            os.remove (local_file_with_date)
        else:
            logger.error("The file does not exist")

        # Delay execution for 1 second
        time.sleep(1)

        # Remove the handlers and close the log file
        handlers = logger.handlers[:]
        for handler in handlers:
            handler.close()
            logger.removeHandler(handler)

        # Open the log file
        os.system('start logfile.log')
    except Exception as e:
        logging.error(f"Script exited with error: {str(e)}")
